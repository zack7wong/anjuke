from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors
from openpyxl.styles import Fill,fills
from openpyxl.formatting.rule import ColorScaleRule
#打开一个workbook

wb = load_workbook("/home/python/Desktop/厦门_学校.xlsx")
sheet = wb.active

# 获得最大列和最大行
# print(sheet.max_row)
# print(sheet.max_column)

# for cell in list(sheet.columns)[1]:
#     print(cell.value)
for i in range(2, sheet.max_row+1):
    for j in range(1, 2):
        print(sheet.cell(row=i,column=j).value)
###### #########
# for i in range(2, sheet.max_row+1):
#     for j in range(1, 3):
#         print(sheet.cell(row=i,column=1).value)

######    读取A1全部行     #########
# for row_cell in sheet['A1':'A'+str(sheet.max_row)]:
#     for cell in row_cell:
#         print(cell.value)