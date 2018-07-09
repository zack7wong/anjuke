# -*- coding: utf-8 -*-
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import random
import xlwt
import time

from openpyxl import Workbook
import tldextract
from company.spiders.companyspider import CompanyspiderSpider


class CompanyPipeline(object):
    def __init__(self):
        self.file = xlwt.Workbook(encoding='utf8')
        self.table1 = self.file.add_sheet('sheet name', cell_overwrite_ok=True)
        for i in range(0, CompanyspiderSpider.nrows):
            for j in range(0, CompanyspiderSpider.ncols):
                # print(CompanyspiderSpider.table.cell(i, j).value)
                self.table1.write(i, j, CompanyspiderSpider.table.cell(i, j).value)  # 全部复制
                # file.save('demo1.xls')
        self.name = CompanyspiderSpider.excel.split('.')[0]+'output.xlsx'

    def process_item(self, item, spider):
        # 关键词123: title,keywords,excess
        # print('初始item',item)
        item['url'] = '.'.join(tldextract.extract(item.get('url'))[:3])   # url去掉http
        if item.get('unique_words') is not None:
            unique_words_list = []
            for i in item.get('unique_words'):
                unique_words_list.append(i)
            # unique_words_list = list(item.get('unique_words'))  # 含有关键词的词
        try:
            print('unique_words_list',unique_words_list)
            if len(unique_words_list) > 1:  # 至少两个
                if item.get('keywords2') is None:
                    item['keywords2'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]
                    # item['keywords2'] = unique_words_list[0]

                if item.get('keywords2') == item.get('keywords1'):
                    item['keywords2'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]
                    unique_words_list.pop(unique_words_list.index(item['keywords2']))
                    if item.get('keywords2') == item.get('keywords1'):
                        item['keywords2'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]
                        unique_words_list.pop(unique_words_list.index(item['keywords2']))

                if item.get('keywords3') is None:
                    item['keywords3'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]
                    unique_words_list.pop(unique_words_list.index(item['keywords2']))

                    if item.get('keywords3') == item.get('keywords2'):
                        item['keywords3'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]

            # if item['description'] is None:
            # item['excess'] = unique_words_list[random.randint(1,len(unique_words_list))]
        except Exception as e:
            print(e)

        print('处理后item>>', item)
        if item.get('keywords1') is not None:
            if item.get('keywords3') is not None and item.get('keywords2') is not None:
                try:
                    list = [item.get('keywords1'), item.get('keywords2'), item.get('keywords3')]
                    list2 = sorted(list, key=lambda x: len(x), reverse=True)
                    self.table1.write(int(item['url_index_doc']) + 1, 9, list2[0])
                    self.table1.write(int(item['url_index_doc']) + 1, 10, list2[1])
                    self.table1.write(int(item['url_index_doc']) + 1, 11, list2[2])
                    # self.table1.write(int(item['url_index_doc']) + 1, 12, item.get('copyrt'))
                    self.file.save(self.name)
                    print(" save ok! *******************")
                except Exception as e:
                    print(e)
                return item

        # if item.get('keywords3') is not None and item.get('keywords2') is not None:
        #     try:
        #         list = [item.get('keywords1'), item.get('keywords2'), item.get('keywords3')]
        #         list2 = sorted(list, key=lambda x: len(x), reverse=True)
        #
        #         line = [item.get('title'), list2[0], list2[1], list2[2], item.get('copyrt'), item.get('url')]
        #         self.ws.append(line)  # 将数据以行的形式添加到xlsx中
        #         # self.wb.save('/home/python/Desktop/1.xlsx')
        #         self.wb.save('output.xlsx')
        #         print(" save ok!")
        #     except Exception as e:
        #         print(e)
        # return item
