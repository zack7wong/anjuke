# -*- coding: utf-8 -*-
# Define your item pipelines here
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import xlwt
import random
from openpyxl import Workbook
from school.spiders.schoolspider import SchoolspiderSpider


class SchoolPipeline(object):
    def __init__(self):
        self.file = xlwt.Workbook()
        self.table1 = self.file.add_sheet('sheet name', cell_overwrite_ok=True)
        for i in range(0, SchoolspiderSpider.nrows):
            for j in range(0, SchoolspiderSpider.ncols):
                # print(SchoolspiderSpider.table.cell(i, j).value)
                self.table1.write(i, j, SchoolspiderSpider.table.cell(i, j).value)  # 复制
        # file.save('demo1.xls')

    def process_item(self,item,spider):
        # 关键词123: title,keywords,excess
        print('初始item', item)

        if item.get('unique_words') is not None:
            unique_words_list = []
            for i in item.get('unique_words'):
                unique_words_list.append(i)
            # unique_words_list = list(item.get('unique_words'))  # 含有关键词的词
            try:
                # print('unique_words_list-->', unique_words_list)
                if len(unique_words_list) > 1:
                    if item.get('keywords1') is None:
                        item['keywords1']= item['title']
                        # print(item['title'],'<<item["title"<')

                    if item.get('keywords2') is None:
                        if item.get('title') != item.get('keywords1'):
                            item['keywords2'] = item.get('title')
                        else:
                            item['keywords2'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]
                    if item.get('keywords2') == item.get('keywords1'):
                        item['keywords2'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]
                        unique_words_list.pop(unique_words_list.index(item['keywords2']))
                        if item.get('keywords2') == item.get('keywords1'):
                            item['keywords2'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]
                            unique_words_list.pop(unique_words_list.index(item['keywords2']))

                    if item.get('keywords3') is None:
                        item['keywords3'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]
                        unique_words_list.pop(unique_words_list.index(item['keywords3']))

                        if item.get('keywords3') == item.get('keywords2'):
                            item['keywords3'] = unique_words_list[random.randint(0, len(unique_words_list)-1)]
                # if item['description'] is None:
                # item['excess'] = unique_words_list[random.randint(1,5)]
            except Exception as e:
                print(e)
            else:
                # print('\n')
                print('处理后item>>', item)
                if item.get('keywords1') is not None:
                    if item.get('keywords3') is not None and item.get('keywords2') is not None:
                        try:
                            list = [item.get('keywords1'), item.get('keywords2'), item.get('keywords3')]
                            list2 = sorted(list, key=lambda x: len(x),reverse=True)

                            self.table1.write(int(item['url_index_doc'])+1, 9, list2[0])
                            self.table1.write(int(item['url_index_doc'])+1, 10, list2[1])
                            self.table1.write(int(item['url_index_doc'])+1, 11, list2[2])
                            self.table1.write(int(item['url_index_doc'])+1, 12, item.get('copyrt'))
                            self.file.save('output.xlsx')
                            # print(" save ok! *******************")
                        except Exception as e:
                            print(e)
                        return item

                # self.table1.write(int(item['url_index_doc'])+1, 9, item.get('keywords1'))
                # self.table1.write(int(item['url_index_doc'])+1, 10, item.get('keywords2'))
                # self.table1.write(int(item['url_index_doc'])+1, 11, item.get('keywords3'))
                # self.file.save('output.xlsx')
                # print(" save ok!")
                # return item
        # ++++++++++++++++++++++++++++++++++++++++++++++++++++

        # print('处理后item>>', item)
        # #
        # if item.get('keywords3') is not None and item.get('keywords2') is not None:
        #     try:
        #         list = [item.get('keywords1'), item.get('keywords2'), item.get('keywords3')]
        #         list2 = sorted(list, key=lambda x: len(x),reverse=True)
        #
        #         line =[item.get('title'), list2[0], list2[1], list2[2], item.get('copyrt'), item.get('url')]
        #         self.ws.append(line)  # 将数据以行的形式添加到xlsx中
        #         # self.wb.save('/home/python/Desktop/1.xlsx')
        #         self.wb.save('output.xlsx')
        #         print(" save ok!")
        #     except Exception as e:
        #         print(e)
        # #
        # # else:
        # #     try:
        # #         line2 =[item.get('title'),item.get('url')]
        # #         self.ws2.append(line2)  # 将数据以行的形式添加到xlsx中
        # #         self.wb2.save('/home/python/Desktop/2.xlsx')
        # #         print(" save ok!")
        # #     except Exception as e:
        # #         print(e)
        # return item
#$ ============================================

    # def process_item(self, item, spider):
    #     # 关键词123: title,keywords,excess
    #     # print('初始item',item)
    #     if item.get('keywords1') is not None:
    #         try:
    #             line = [item.get('title'), item.get('keywords1'), item.get('keywords2'), item.get('keywords3'),
    #                     item.get('copyrt'), item.get('url')]
    #             self.ws.append(line)  # 将数据以行的形式添加到xlsx中
    #             self.wb.save('/home/python/Desktop/1.xlsx')
    #             print(" save1 ok!")
    #         except Exception as e:
    #             print(e)
    #     else:
    #         try:
    #             line2 =[item.get('title'),item.get('url')]
    #             self.ws2.append(line2)  # 将数据以行的形式添加到xlsx中
    #             self.wb.save('/home/python/Desktop/2.xlsx')
    #             print(" save2 ok!")
    #         except Exception as e:
    #             print(e)
    #
            # return item
